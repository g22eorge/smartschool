from django.contrib.auth.decorators import login_required
from django.shortcuts import render, get_object_or_404
from django.utils import timezone
from django.http import HttpResponse
from django.contrib import messages
import openpyxl
from io import BytesIO
from datetime import datetime

from ..models import QRCode, Attendance, Module

@login_required
def attendance_report(request, module_id):
    module = get_object_or_404(Module, id=module_id)
    
    # Check if user has permission to view this report
    if not request.user.is_superuser and request.user not in module.lecturers.all():
        messages.error(request, 'You do not have permission to view this report.')
        return redirect('attendance:dashboard')
    
    # Get all QR codes for this module
    qr_codes = QRCode.objects.filter(module=module).order_by('-created_at')
    
    # Get all students enrolled in this module
    students = module.students.all()
    
    # Calculate attendance stats
    attendance_stats = []
    for student in students:
        total_sessions = qr_codes.count()
        attended_sessions = Attendance.objects.filter(
            student=student,
            qrcode__in=qr_codes,
            status='present'
        ).count()
        
        attendance_rate = (attended_sessions / total_sessions * 100) if total_sessions > 0 else 0
        
        attendance_stats.append({
            'student': student,
            'total_sessions': total_sessions,
            'attended_sessions': attended_sessions,
            'attendance_rate': round(attendance_rate, 2)
        })
    
    context = {
        'module': module,
        'attendance_stats': attendance_stats,
        'total_sessions': qr_codes.count(),
    }
    
    return render(request, 'attendance/attendance_report.html', context)

@login_required
def download_attendance(request, qrcode_id):
    qrcode = get_object_or_404(QRCode, id=qrcode_id)
    
    # Check if user has permission to download this report
    if not request.user.is_superuser and request.user != qrcode.lecturer:
        messages.error(request, 'You do not have permission to download this report.')
        return redirect('attendance:attendance_report', module_id=qrcode.module.id)
    
    # Create a new workbook and add a worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Attendance - {qrcode.module.code}"
    
    # Add headers
    headers = ['Student ID', 'Name', 'Email', 'Status', 'Timestamp']
    for col_num, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_num, value=header)
    
    # Add data
    attendances = Attendance.objects.filter(qrcode=qrcode).select_related('student')
    for row_num, attendance in enumerate(attendances, 2):
        ws.cell(row=row_num, column=1, value=attendance.student.student_id)
        ws.cell(row=row_num, column=2, value=attendance.student.get_full_name())
        ws.cell(row=row_num, column=3, value=attendance.student.email)
        ws.cell(row=row_num, column=4, value=attendance.get_status_display())
        ws.cell(row=row_num, column=5, value=attendance.timestamp.strftime('%Y-%m-%d %H:%M:%S'))
    
    # Create a response with the Excel file
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename=attendance_{qrcode.module.code}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    # Save the workbook to the response
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response
